import sys
import os
import logging
import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# CONFIG─────────────────────────────────────────────

TICKERS     = ["^GSPC", "AAPL", "MSFT", "NVDA", "AMD", "TSLA"]
START_DATE  = "2025-01-01"
END_DATE    = "2026-06-30"
INTERVAL    = "1mo"
OUTPUT_FILE = "Financial_Data_Report.xlsx"


# LOGGING─────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    datefmt="%H:%M:%S",)
log = logging.getLogger(__name__)


# 1. DOWNLOAD─────────────────────────────────────────────

log.info("Downloading market data for %s ...", TICKERS)

try:
    data = yf.download(
        tickers=TICKERS,
        start=START_DATE,
        end=END_DATE,
        interval=INTERVAL,
        auto_adjust=False,
        group_by="ticker",
        progress=False,)
except Exception as e:
    log.error("Download failed: %s", e)
    sys.exit(1)

ticker_dfs: dict[str, pd.DataFrame] = {}

for ticker in TICKERS:
    try:
        ticker_dfs[ticker] = data[ticker][["Open", "High", "Low", "Close", "Volume"]].copy()
    except KeyError:
        log.warning("Ticker %s not found in download result — skipping.", ticker)

if not ticker_dfs:
    log.error("No data retrieved. Exiting.")
    sys.exit(1)


# 2. CLEANING─────────────────────────────────────────────

log.info("Cleaning data ...")

def clean_ohlcv(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # 1. Standardise date index
    df.index = pd.to_datetime(df.index)
    df = df.sort_index()

    # 2. Remove duplicates
    df = df[~df.index.duplicated(keep="first")]

    # 3. Lowercase column names
    df.columns = [c.lower() for c in df.columns]

    # 4. Handle missing values
    df = df.ffill().bfill()

    # 5. OHLC relationship validation
    df["price_error"] = (
        (df["high"] < df["open"])  |
        (df["high"] < df["close"]) |
        (df["low"]  > df["open"])  |
        (df["low"]  > df["close"]) |
        (df["high"] < df["low"])   |
        (df["volume"] < 0))

    # 6. Outlier detection (z-score > 3)
    for col in ["open", "high", "low", "close", "volume"]:
        if col in df.columns:
            z = (df[col] - df[col].mean()) / df[col].std()
            df[f"{col}_outlier"] = np.abs(z) > 3

    return df


def enrich(df: pd.DataFrame, column: str = "close") -> pd.DataFrame:
    df = df.copy()
    if df.empty:
        return df

    base = df[column].iloc[0]
    df[f"{column}_norm"]  = (df[column] / base) * 100
    df["return"]          = df["close"].pct_change()
    df["log_return"]      = np.log(df["close"] / df["close"].shift(1))
    df["MA_3"]            = df["close"].rolling(window=3).mean()
    df["MA_6"]            = df["close"].rolling(window=6).mean()
    return df


def full_pipeline(df: pd.DataFrame) -> pd.DataFrame:
    return enrich(clean_ohlcv(df))


ticker_dfs_clean: dict[str, pd.DataFrame] = {
    ticker: full_pipeline(df) for ticker, df in ticker_dfs.items()}

# Aligned normalised close (used for cross-asset comparison)
aligned = (
    pd.concat([df["close_norm"] for df in ticker_dfs_clean.values()], axis=1)
    .set_axis(list(ticker_dfs_clean.keys()), axis=1)
    .ffill())


# 3. ANALYTICS─────────────────────────────────────────────

log.info("Running analytics ...")

RISK_FREE_RATE = 0.05          # annualised — approximate US risk-free rate
PERIODS_PER_YEAR = 12          # monthly data

def compute_analytics(ticker_dfs_clean: dict, rf: float, periods: int) -> dict:
    """
    Returns a dict of analytics DataFrames:
      - rolling_sharpe  : rolling 6-month Sharpe ratio per ticker
      - drawdown        : drawdown series per ticker
      - momentum        : latest momentum signals per ticker
      - correlation     : return correlation matrix
      - risk_return     : annualised risk/return summary per ticker
                          (includes beta and information ratio vs ^GSPC)
    """
    # Align returns
    returns = pd.concat(
        [df["return"] for df in ticker_dfs_clean.values()],
        axis=1
    ).set_axis(list(ticker_dfs_clean.keys()), axis=1)

    # Rolling Sharpe (6-month window) ──────────────────────────────────────
    rf_per_period = rf / periods
    excess = returns - rf_per_period

    rolling_sharpe = (
        excess.rolling(6).mean() / returns.rolling(6).std()
    ) * np.sqrt(periods)

    # Drawdown ─────────────────────────────────────────────────────────────
    close_prices = pd.concat(
        [df["close"] for df in ticker_dfs_clean.values()],
        axis=1
    ).set_axis(list(ticker_dfs_clean.keys()), axis=1)

    rolling_max = close_prices.cummax()
    drawdown = (close_prices - rolling_max) / rolling_max * 100  # as %

    # Momentum signals ─────────────────────────────────────────────────────
    momentum_rows = []
    for ticker, df in ticker_dfs_clean.items():
        latest = df.dropna(subset=["close", "MA_3", "MA_6"]).iloc[-1]
        above_ma3 = latest["close"] > latest["MA_3"]
        above_ma6 = latest["close"] > latest["MA_6"]

        if above_ma3 and above_ma6:
            signal = "BULLISH"
        elif not above_ma3 and not above_ma6:
            signal = "BEARISH"
        else:
            signal = "MIXED"

        momentum_rows.append({
            "Ticker":        ticker,
            "Latest Close":  round(latest["close"], 2),
            "MA_3":          round(latest["MA_3"],  2),
            "MA_6":          round(latest["MA_6"],  2),
            "Above MA3":     above_ma3,
            "Above MA6":     above_ma6,
            "Signal":        signal,})

    momentum = pd.DataFrame(momentum_rows).set_index("Ticker")

    # Correlation matrix ────────────────────────────────────────────────────
    correlation = returns.corr().round(4)

    # Risk / Return summary ─────────────────────────────────────────────────
    ann_return = returns.mean() * periods
    ann_vol    = returns.std()  * np.sqrt(periods)
    sharpe     = (ann_return - rf) / ann_vol
    max_dd     = drawdown.min()

    # Beta vs ^GSPC ─────────────────────────────────────────────────────────
    # Beta = Cov(asset, benchmark) / Var(benchmark)
    # Uses full-period returns; ^GSPC must be in the ticker set
    betas = {}
    info_ratios = {}

    if "^GSPC" in returns.columns:
        bench = returns["^GSPC"].dropna()
        bench_ann_return = bench.mean() * periods

        for ticker in returns.columns:
            asset = returns[ticker].dropna()
            # Align on common dates
            aligned_pair = pd.concat([asset, bench], axis=1).dropna()
            aligned_pair.columns = ["asset", "bench"]

            if len(aligned_pair) < 3:
                betas[ticker]       = np.nan
                info_ratios[ticker] = np.nan
                continue

            cov_matrix  = np.cov(aligned_pair["asset"], aligned_pair["bench"])
            beta        = cov_matrix[0, 1] / cov_matrix[1, 1]
            betas[ticker] = round(beta, 3)

            # Information Ratio = (asset return - benchmark return) / tracking error
            # Tracking error = std dev of (asset returns - benchmark returns)
            active_returns  = aligned_pair["asset"] - aligned_pair["bench"]
            tracking_error  = active_returns.std() * np.sqrt(periods)
            active_ann      = active_returns.mean() * periods
            info_ratios[ticker] = round(
                active_ann / tracking_error if tracking_error != 0 else np.nan, 3)
    else:
        log.warning("^GSPC not in tickers — beta and IR skipped.")
        for ticker in returns.columns:
            betas[ticker]       = np.nan
            info_ratios[ticker] = np.nan

    risk_return = pd.DataFrame({
        "Annualised Return (%)": (ann_return * 100).round(2),
        "Annualised Vol (%)":    (ann_vol    * 100).round(2),
        "Sharpe Ratio":          sharpe.round(3),
        "Beta (vs S&P 500)":     pd.Series(betas),
        "Information Ratio":     pd.Series(info_ratios),
        "Max Drawdown (%)":      max_dd.round(2),})

    return {
        "rolling_sharpe": rolling_sharpe.round(3),
        "drawdown":        drawdown.round(2),
        "momentum":        momentum,
        "correlation":     correlation,
        "risk_return":     risk_return,}


analytics = compute_analytics(ticker_dfs_clean, RISK_FREE_RATE, PERIODS_PER_YEAR)


# 4. EXCEL OUTPUT─────────────────────────────────────────────

log.info("Building workbook ...")

# Always start from a fresh workbook — loading an existing file risks
# carrying forward stale chart/drawing XML which causes Excel repair warnings.
wb = Workbook()

def get_or_create_sheet(wb, name, first=False):
    """Return a cleared sheet, creating it if necessary."""
    if name in wb.sheetnames:
        ws = wb[name]
        ws.delete_rows(1, ws.max_row)
    elif first:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(name)
    return ws


# Sheet 1: Raw Data ──────────────────────────────────────────────────────

ws_raw = get_or_create_sheet(wb, "Raw Data", first=True)
row = 1

for ticker, df in ticker_dfs.items():
    ws_raw.cell(row=row, column=1).value = ticker
    row += 1
    for col_num, col in enumerate(["Date"] + list(df.columns), start=1):
        ws_raw.cell(row=row, column=col_num).value = col
    row += 1
    for idx, values in df.iterrows():
        ws_raw.cell(row=row, column=1).value = idx.strftime("%Y-%m-%d")
        for col_num, value in enumerate(values, start=2):
            ws_raw.cell(row=row, column=col_num).value = value
        row += 1
    row += 2


# Sheet 2: Cleaned Data ─────────────────────────────────────────────────

ws_clean = get_or_create_sheet(wb, "Cleaned Data")
row = 1

for ticker, df in ticker_dfs_clean.items():
    ws_clean.cell(row=row, column=1).value = ticker
    row += 1
    headers = ["Date"] + list(df.columns)
    for col_num, header in enumerate(headers, start=1):
        ws_clean.cell(row=row, column=col_num).value = header
    row += 1
    for idx, values in df.iterrows():
        ws_clean.cell(row=row, column=1).value = idx.strftime("%Y-%m-%d")
        for col_num, value in enumerate(values, start=2):
            ws_clean.cell(row=row, column=col_num).value = value
        row += 1
    row += 2


# Sheet 3: Summary Stats ────────────────────────────────────────────────

ws_summary = get_or_create_sheet(wb, "Summary Stats")

for c, h in enumerate(["Ticker", "Mean Close", "Std Dev Close",
                        "Min Close", "Max Close", "% Missing Filled"], start=1):
    ws_summary.cell(row=1, column=c).value = h

for row_num, ticker in enumerate(ticker_dfs, start=2):
    raw   = ticker_dfs[ticker]
    clean = ticker_dfs_clean[ticker]
    miss_before = raw.isna().sum().sum()
    miss_after  = clean[["open","high","low","close","volume"]].isna().sum().sum()
    pct_filled  = 0 if miss_before == 0 else (miss_before - miss_after) / miss_before * 100

    ws_summary.cell(row=row_num, column=1).value = ticker
    ws_summary.cell(row=row_num, column=2).value = round(clean["close"].mean(), 2)
    ws_summary.cell(row=row_num, column=3).value = round(clean["close"].std(),  2)
    ws_summary.cell(row=row_num, column=4).value = round(clean["close"].min(),  2)
    ws_summary.cell(row=row_num, column=5).value = round(clean["close"].max(),  2)
    ws_summary.cell(row=row_num, column=6).value = round(pct_filled, 2)


# Sheet 4: Change Log ───────────────────────────────────────────────────

ws_log = get_or_create_sheet(wb, "Change Log")

for c, h in enumerate(["Ticker", "Duplicate Rows Removed", "Missing Values Before",
                        "Missing Values After", "Outliers Detected"], start=1):
    ws_log.cell(row=1, column=c).value = h

for row_num, ticker in enumerate(ticker_dfs, start=2):
    raw   = ticker_dfs[ticker]
    clean = ticker_dfs_clean[ticker]

    ws_log.cell(row=row_num, column=1).value = ticker
    ws_log.cell(row=row_num, column=2).value = int(raw.index.duplicated().sum())
    ws_log.cell(row=row_num, column=3).value = int(raw.isna().sum().sum())
    ws_log.cell(row=row_num, column=4).value = int(clean[["open","high","low","close","volume"]].isna().sum().sum())
    ws_log.cell(row=row_num, column=5).value = int(clean.filter(like="_outlier").sum().sum())


# Sheet 5: Analytics ───────────────────────────────────────────────────

ws_analytics = get_or_create_sheet(wb, "Analytics")
r = 1

def write_section(ws, title: str, df: pd.DataFrame, start_row: int) -> int:
    """Write a titled table to the sheet. Returns the next free row."""
    ws.cell(row=start_row, column=1).value = title
    start_row += 1

    # Header — index label + column names
    ws.cell(row=start_row, column=1).value = df.index.name or "Ticker"
    for c, col in enumerate(df.columns, start=2):
        ws.cell(row=start_row, column=c).value = col
    start_row += 1

    # Data rows
    for idx, row_data in df.iterrows():
        ws.cell(row=start_row, column=1).value = (
            idx.strftime("%Y-%m-%d") if hasattr(idx, "strftime") else str(idx)
        )
        for c, val in enumerate(row_data, start=2):
            ws.cell(row=start_row, column=c).value = val
        start_row += 1

    return start_row + 2   # blank gap between sections


# Risk / Return summary
r = write_section(ws_analytics, "── Risk & Return Summary (annualised)", analytics["risk_return"], r)

# Momentum signals
r = write_section(ws_analytics, "── Momentum Signals (latest period)", analytics["momentum"], r)

# Correlation matrix
r = write_section(ws_analytics, "── Return Correlation Matrix", analytics["correlation"], r)

# Rolling Sharpe
rolling_sharpe_out = analytics["rolling_sharpe"].copy()
rolling_sharpe_out.index.name = "Date"
r = write_section(ws_analytics, "── Rolling 6-Month Sharpe Ratio", rolling_sharpe_out, r)

# Drawdown
drawdown_out = analytics["drawdown"].copy()
drawdown_out.index.name = "Date"
r = write_section(ws_analytics, "── Drawdown (% from peak)", drawdown_out, r)


# Sheet 6: Charts data table (openpyxl) + Charts file (xlsxwriter) ────
# openpyxl's chart XML causes Excel repair warnings on Mac/Windows.
# Strategy: write chart DATA into the main workbook (Charts sheet),
# and write the actual charts into a companion _charts.xlsx via xlsxwriter
# which produces clean, warning-free XML.

tickers_list = list(ticker_dfs_clean.keys())
dates_index  = ticker_dfs_clean[tickers_list[0]].index
n_dates      = len(dates_index)
dd_df        = analytics["drawdown"]

# Data-only Charts sheet in main workbook ───────────────────────────────

if "Charts" in wb.sheetnames:
    del wb["Charts"]
ws_charts = wb.create_sheet("Charts")

ws_charts.cell(row=1, column=1).value = "── Normalised Price Data (Base = 100)"
ws_charts.cell(row=2, column=1).value = "Date"
for c, ticker in enumerate(tickers_list, start=2):
    ws_charts.cell(row=2, column=c).value = ticker

for r, date in enumerate(dates_index, start=3):
    ws_charts.cell(row=r, column=1).value = date.strftime("%Y-%m-%d")
    for c, ticker in enumerate(tickers_list, start=2):
        val = ticker_dfs_clean[ticker]["close_norm"].get(date, None)
        ws_charts.cell(row=r, column=c).value = (
            round(float(val), 2) if val is not None and not pd.isna(val) else None)

norm_end_row = 2 + n_dates
dd_hdr_row   = norm_end_row + 3

ws_charts.cell(row=dd_hdr_row,     column=1).value = "── Drawdown Data (% from peak)"
ws_charts.cell(row=dd_hdr_row + 1, column=1).value = "Date"
for c, ticker in enumerate(tickers_list, start=2):
    ws_charts.cell(row=dd_hdr_row + 1, column=c).value = ticker

for r, date in enumerate(dates_index, start=dd_hdr_row + 2):
    ws_charts.cell(row=r, column=1).value = date.strftime("%Y-%m-%d")
    for c, ticker in enumerate(tickers_list, start=2):
        val = dd_df[ticker].get(date, None)
        ws_charts.cell(row=r, column=c).value = (
            round(float(val), 2) if val is not None and not pd.isna(val) else None)

# Companion charts file via xlsxwriter ─────────────────────────────────
import xlsxwriter

CHARTS_FILE = OUTPUT_FILE.replace(".xlsx", "_charts.xlsx")
xw  = xlsxwriter.Workbook(CHARTS_FILE)
xws = xw.add_worksheet("Charts")

# Write normalised price data (0-indexed)

xws.write(0, 0, "Date")
for c, ticker in enumerate(tickers_list, start=1):
    xws.write(0, c, ticker)
for r, date in enumerate(dates_index):
    xws.write(r + 1, 0, date.strftime("%Y-%m-%d"))
    for c, ticker in enumerate(tickers_list, start=1):
        val = ticker_dfs_clean[ticker]["close_norm"].get(date, None)
        if val is not None and not pd.isna(val):
            xws.write(r + 1, c, round(float(val), 2))

norm_last = n_dates   # last data row, 0-indexed

# Write drawdown data
dd_hdr = norm_last + 3
xws.write(dd_hdr, 0, "Date")
for c, ticker in enumerate(tickers_list, start=1):
    xws.write(dd_hdr, c, ticker)
for r, date in enumerate(dates_index):
    xws.write(dd_hdr + 1 + r, 0, date.strftime("%Y-%m-%d"))
    for c, ticker in enumerate(tickers_list, start=1):
        val = dd_df[ticker].get(date, None)
        if val is not None and not pd.isna(val):
            xws.write(dd_hdr + 1 + r, c, round(float(val), 2))

dd_last = dd_hdr + n_dates
chart_col = len(tickers_list) + 2   # place charts to the right of data

# Chart 1 — Normalised price
c1 = xw.add_chart({"type": "line"})
c1.set_title({"name": "Normalised Price Trend (Base = 100)"})
c1.set_x_axis({"name": "Date"})
c1.set_y_axis({"name": "Price (Indexed to 100)"})
c1.set_style(10)
c1.set_size({"width": 700, "height": 380})
for c, ticker in enumerate(tickers_list, start=1):
    c1.add_series({
        "name":       ["Charts", 0, c],
        "categories": ["Charts", 1, 0, norm_last, 0],
        "values":     ["Charts", 1, c, norm_last, c],
        "line":       {"width": 1.5},})
xws.insert_chart(0, chart_col, c1)

# Chart 2 — Drawdown
c2 = xw.add_chart({"type": "line"})
c2.set_title({"name": "Drawdown from Peak (%)"})
c2.set_x_axis({"name": "Date"})
c2.set_y_axis({"name": "Drawdown (%)"})
c2.set_style(10)
c2.set_size({"width": 700, "height": 380})
for c, ticker in enumerate(tickers_list, start=1):
    c2.add_series({
        "name":       ["Charts", dd_hdr, c],
        "categories": ["Charts", dd_hdr + 1, 0, dd_last, 0],
        "values":     ["Charts", dd_hdr + 1, c, dd_last, c],
        "line":       {"width": 1.5},})
xws.insert_chart(dd_hdr, chart_col, c2)

xw.close()
log.info("Charts saved → %s", CHARTS_FILE)


# 5. FORMATTING─────────────────────────────────────────────

def format_workbook(wb: Workbook) -> None:
    header_fill = PatternFill(fill_type="solid", start_color="4F81BD", end_color="4F81BD")
    header_font = Font(bold=True, color="FFFFFF")
    thin         = Side(style="thin")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    red_fill     = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
    grey_fill    = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")

    for ws in wb.worksheets:
        # Charts sheet contains chart objects — skip formatting passes that
        # iterate cells, as openpyxl raises errors on chart-only sheets
        if ws.title == "Charts":
            ws.sheet_properties.tabColor = "FF8C00"
            continue

        ws.freeze_panes = "A3" if ws.title in ("Raw Data", "Cleaned Data") else "A2"

        # Track which rows are header rows so we can skip alternate shading on them
        header_rows = set()

        for row in ws.iter_rows():
            values = [cell.value for cell in row]
            is_header = "Date" in values or "Ticker" in values
            if is_header:
                header_rows.add(row[0].row)
                for cell in row:
                    cell.font      = header_font
                    cell.fill      = header_fill
                    cell.alignment = Alignment(horizontal="center")

            for cell in row:
                # FIX 2: highlight True booleans (price errors / outliers)
                if cell.value is True:
                    cell.fill = red_fill

        # FIX 3: alternate shading — skip header rows and already-filled cells
        for r_idx in range(1, ws.max_row + 1):
            if r_idx in header_rows:
                continue
            if r_idx % 2 == 0:
                for cell in ws[r_idx]:
                    # Only shade if the cell has no fill already set
                    if cell.fill.fill_type in (None, "none"):
                        cell.fill = grey_fill

        # Borders + auto-width
        for col in ws.columns:
            max_len = 0
            letter  = get_column_letter(col[0].column)
            for cell in col:
                cell.border = border
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = max_len + 3

        ws.auto_filter.ref = ws.dimensions

        # Number formatting ──────────────────────────────────────────────
        # Read header row to map column letter → column name
        header_row = None
        for row in ws.iter_rows(max_row=20):
            if any(cell.value == "Date" for cell in row):
                header_row = row
                break

        if header_row:
            col_formats = {}
            for cell in header_row:
                name = str(cell.value or "").lower()
                letter = get_column_letter(cell.column)
                if name in ("open", "high", "low", "close",
                            "close_norm", "ma_3", "ma_6",
                            "mean close", "std dev close",
                            "min close", "max close"):
                    col_formats[letter] = '#,##0.00'
                elif name == "volume":
                    col_formats[letter] = '#,##0'
                elif name in ("return", "log_return"):
                    col_formats[letter] = '0.0000'
                elif name == "% missing filled":
                    col_formats[letter] = '0.00'

            # Apply formats to every data cell in that column
            for r in ws.iter_rows(min_row=header_row[0].row + 1):
                for cell in r:
                    letter = get_column_letter(cell.column)
                    if letter in col_formats and isinstance(cell.value, (int, float)):
                        cell.number_format = col_formats[letter]

    wb["Raw Data"].sheet_properties.tabColor      = "808080"
    wb["Cleaned Data"].sheet_properties.tabColor  = "00B050"
    wb["Summary Stats"].sheet_properties.tabColor = "4F81BD"
    wb["Change Log"].sheet_properties.tabColor    = "C0504D"
    wb["Analytics"].sheet_properties.tabColor     = "7030A0"


format_workbook(wb)
wb["Charts"].sheet_properties.tabColor = "FF8C00"

# 6. SAVE─────────────────────────────────────────────

wb.save(OUTPUT_FILE)
log.info("Saved → %s", OUTPUT_FILE)
log.info("Open %s for interactive charts (no repair warnings)", OUTPUT_FILE.replace('.xlsx', '_charts.xlsx'))
